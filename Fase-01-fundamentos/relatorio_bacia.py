# relatorio_bacia.py
# Relatório de caracterização de bacia hidrográfica
# Exporta relatório formatado em .xlsx

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from calc_bacia_hidrografica import (ler_perfil_csv, calcular_forma_bacia,
                                     calcular_declividades, calcular_tempo_concentracao)

EQUACOES = ["kirpich", "ven_te_chow", "johnstone", "corps_engineers", "picking", "carter"]
DECLIVIDADES = ["s1", "s2", "s3"]
COR_TITULO    = "1F4E79"
COR_SECAO     = "2E75B6"
COR_CABECALHO = "BDD7EE"
COR_LINHA_ALT = "DEEAF1"
COR_BRANCO    = "FFFFFF"

def _fonte(bold=False, branco=False, tamanho=10):
    cor = "FFFFFF" if branco else "000000"
    return Font(name="Arial", bold=bold, color=cor, size=tamanho)

def _fundo(cor):
    return PatternFill("solid", fgColor=cor)

def _borda():
    lado = Side(style="thin", color="B8CCE4")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _centro():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _esquerda():
    return Alignment(horizontal="left", vertical="center")


def _titulo_secao(ws, linha, texto):
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=7)
    c = ws.cell(linha, 1, texto)
    c.font = _fonte(bold=True, branco=True, tamanho=11)
    c.fill = _fundo(COR_SECAO)
    c.alignment = _centro()
    c.border = _borda()
    return linha + 1


def _linha_dado(ws, linha, rotulo, valor, complemento=""):
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=2)
    c1 = ws.cell(linha, 1, rotulo)
    c1.font = _fonte(bold=True)
    c1.fill = _fundo(COR_CABECALHO)
    c1.alignment = _esquerda()
    c1.border = _borda()

    ws.merge_cells(start_row=linha, start_column=3, end_row=linha, end_column=5)
    c2 = ws.cell(linha, 3, valor)
    c2.font = _fonte()
    c2.fill = _fundo(COR_BRANCO)
    c2.alignment = _esquerda()
    c2.border = _borda()

    ws.merge_cells(start_row=linha, start_column=6, end_row=linha, end_column=7)
    c3 = ws.cell(linha, 6, complemento)
    c3.font = _fonte()
    c3.fill = _fundo(COR_BRANCO)
    c3.alignment = _esquerda()
    c3.border = _borda()
    return linha + 1


def gerar_xlsx(nome_bacia, area_km2, perimetro_km, lc_km,
               forma, decliv, resultados_tc, classe_kf, classe_kc, nome_arquivo):

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    ws.column_dimensions["A"].width = 26
    for col in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 16

    linha = 1

    # Título principal
    ws.row_dimensions[linha].height = 28
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=7)
    c = ws.cell(linha, 1, "RELATÓRIO DE CARACTERIZAÇÃO DE BACIA HIDROGRÁFICA")
    c.font = _fonte(bold=True, branco=True, tamanho=13)
    c.fill = _fundo(COR_TITULO)
    c.alignment = _centro()
    linha += 1

    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=7)
    c = ws.cell(linha, 1, f"Bacia: {nome_bacia}")
    c.font = _fonte(bold=True, branco=True, tamanho=11)
    c.fill = _fundo(COR_TITULO)
    c.alignment = _centro()
    linha += 2

    # Dados de entrada
    linha = _titulo_secao(ws, linha, "DADOS DE ENTRADA")
    linha = _linha_dado(ws, linha, "Área",                 area_km2,     "km²")
    linha = _linha_dado(ws, linha, "Perímetro",            perimetro_km, "km")
    linha = _linha_dado(ws, linha, "Comprimento talvegue", lc_km,        "km")
    linha += 1

    # Forma da bacia
    linha = _titulo_secao(ws, linha, "FORMA DA BACIA")
    linha = _linha_dado(ws, linha, "Kf — fator de forma",    round(forma["kf"], 4), classe_kf)
    linha = _linha_dado(ws, linha, "Kc — coef. compacidade", round(forma["kc"], 4), classe_kc)
    linha += 1

    # Declividades
    linha = _titulo_secao(ws, linha, "DECLIVIDADES DO TALVEGUE")
    linha = _linha_dado(ws, linha, "S1 — método direto",    round(decliv["s1"], 6), "m/m")
    linha = _linha_dado(ws, linha, "S2 — área equivalente", round(decliv["s2"], 6), "m/m")
    linha = _linha_dado(ws, linha, "S3 — média harmônica",  round(decliv["s3"], 6), "m/m")
    linha += 1

    # Tabela de Tc em horas
    linha = _titulo_secao(ws, linha, "TEMPO DE CONCENTRAÇÃO  —  Tc (horas)")

    cabecalhos_h = ["Equação", "S1", "S2", "S3"]
    for col, cab in enumerate(cabecalhos_h, 1):
        c = ws.cell(linha, col, cab)
        c.font = _fonte(bold=True)
        c.fill = _fundo(COR_CABECALHO)
        c.alignment = _centro()
        c.border = _borda()
    linha += 1

    for idx, eq in enumerate(EQUACOES):
        cor = COR_LINHA_ALT if idx % 2 == 0 else COR_BRANCO
        valores = [eq,
                   round(resultados_tc["s1"][eq], 4),
                   round(resultados_tc["s2"][eq], 4),
                   round(resultados_tc["s3"][eq], 4)]
        for col, val in enumerate(valores, 1):
            c = ws.cell(linha, col, val)
            c.font = _fonte(bold=(col == 1))
            c.fill = _fundo(cor)
            c.alignment = _esquerda() if col == 1 else _centro()
            c.border = _borda()
        linha += 1

    linha += 1

    # Tabela de Tc em minutos
    linha = _titulo_secao(ws, linha, "TEMPO DE CONCENTRAÇÃO  —  Tc (minutos)")

    cabecalhos_min = ["Equação", "S1", "S2", "S3"]
    for col, cab in enumerate(cabecalhos_min, 1):
        c = ws.cell(linha, col, cab)
        c.font = _fonte(bold=True)
        c.fill = _fundo(COR_CABECALHO)
        c.alignment = _centro()
        c.border = _borda()
    linha += 1

    for idx, eq in enumerate(EQUACOES):
        cor = COR_LINHA_ALT if idx % 2 == 0 else COR_BRANCO
        valores = [eq,
                   round(resultados_tc["s1"][eq] * 60, 2),
                   round(resultados_tc["s2"][eq] * 60, 2),
                   round(resultados_tc["s3"][eq] * 60, 2)]
        for col, val in enumerate(valores, 1):
            c = ws.cell(linha, col, val)
            c.font = _fonte(bold=(col == 1))
            c.fill = _fundo(cor)
            c.alignment = _esquerda() if col == 1 else _centro()
            c.border = _borda()
        linha += 1

    linha += 1
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=7)
    c = ws.cell(linha, 1, "Referência: Silveira (2005), RBRH v.10 n.1, pp.5-23.")
    c.font = Font(name="Arial", size=9, italic=True, color="595959")
    c.alignment = _esquerda()

    wb.save(nome_arquivo)


if __name__ == "__main__":

    print("Relatório de Caracterização de Bacia Hidrográfica")
    print("-" * 50)

    caminho_perfil = input("Caminho do arquivo CSV do perfil do talvegue: ").strip()
    if not os.path.exists(caminho_perfil):
        print("Arquivo não encontrado.")
        exit()

    try:
        area_km2     = float(input("Área da bacia (km²): "))
        perimetro_km = float(input("Perímetro da bacia (km): "))
        lc_km        = float(input("Comprimento do talvegue principal (km): "))
        nome_bacia   = input("Nome da bacia: ").strip()
    except ValueError:
        print("Valor inválido. Use ponto como separador decimal.")
        exit()

    perfil  = ler_perfil_csv(caminho_perfil)
    forma   = calcular_forma_bacia(area_km2, perimetro_km, lc_km)
    decliv  = calcular_declividades(perfil)

    resultados_tc = {s: calcular_tempo_concentracao(lc_km, decliv[s]) for s in DECLIVIDADES}

    if forma["kf"] >= 0.75:   classe_kf = "Alta propensão a grandes enchentes"
    elif forma["kf"] >= 0.50: classe_kf = "Tendência mediana a grandes enchentes"
    else:                      classe_kf = "Não sujeita a grandes enchentes"

    if forma["kc"] <= 1.25:   classe_kc = "Alta propensão a grandes enchentes"
    elif forma["kc"] <= 1.50: classe_kc = "Tendência mediana a grandes enchentes"
    else:                      classe_kc = "Não sujeita a grandes enchentes"

    nome_arquivo = f"relatorio_{nome_bacia.replace(' ', '_')}.xlsx"
    gerar_xlsx(nome_bacia, area_km2, perimetro_km, lc_km,
               forma, decliv, resultados_tc, classe_kf, classe_kc, nome_arquivo)

    print(f"Relatório exportado: {nome_arquivo}")
